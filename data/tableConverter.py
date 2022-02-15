import sys
import re
from openpyxl import load_workbook
from pprint import pprint
from collections import defaultdict

def merge_common(lists): 
    neigh = defaultdict(set) 
    visited = set() 
    for each in lists: 
        for item in each: 
            neigh[item].update(each) 
    def comp(node, neigh = neigh, visited = visited, vis = visited.add): 
        nodes = set([node]) 
        next_node = nodes.pop 
        while nodes: 
            node = next_node() 
            vis(node) 
            nodes |= neigh[node] - visited 
            yield node 
    for node in neigh: 
        if node not in visited: 
            yield sorted(comp(node)) 

def parse_results(results):
    output = {'gl': '-', 'ir': '-', 'ti': '-', 'up': '-'}
    for r in results:
        parts = r.split(' - ')
        if parts[1] == 'GLOOMY':
            output['gl'] = parts[0]
        elif parts[1] == 'IRRITABLE':
            output['ir'] = parts[0]
        elif parts[1] == 'TIMID':
            output['ti'] = parts[0]
        else:
            output['up'] = parts[0]
    return output

def generate_result(result, color):
    if result == "GOOD":
        output = "GOOD</div><div class='symbol'>🎶</div>"
    elif result == "OK":
        output = "OK</div><div class='symbol'>💦</div>"
    elif result == "BAD":
        output = "BAD</div><div class='symbol'>💢</div>"
    elif result == "-":
        output = "-</div><div class='symbol'></div>"
    else:
        output = "{}</div><div class='symbol'></div>".format(result)
    
    class_name = 'result'
    if color:
        class_name = 'result unconfirmed'
    return "\n\t\t<td class='{}'><div class='text'>{}</td>".format(class_name, output)

def create_table(q, tags, top_a, mid_a, bot_a, top_r, mid_r, bot_r, color_top_r, color_mid_r, color_bot_r):
    question = '\n\t<tr>\n\t\t<th colspan="5">{}</th>\n\t</tr>'.format(q)
    header = "\n\t<tr>\n\t\t<td></td>" + "\n\t\t<td class='subheader'>gl<span class='extra'>oomy</span></td>" + "\n\t\t<td class='subheader'>ir<span class='extra'>ritable</span></td>" + "\n\t\t<td class='subheader'>ti<span class='extra'>mid</span></td>" + "\n\t\t<td class='subheader'>up<span class='extra'>beat</span></td>\n\t</tr>"

    top = "\n\t<tr>\n\t\t<td>" + top_a + "</td>" + generate_result(top_r['gl'], color_top_r['gl']) + generate_result(top_r['ir'], color_top_r['ir']) + generate_result(top_r['ti'], color_top_r['ti']) + generate_result(top_r['up'], color_top_r['up']) + "\n\t</tr>"
    mid = "\n\t<tr>\n\t\t<td>" + mid_a + "</td>" + generate_result(mid_r['gl'], color_mid_r['gl']) + generate_result(mid_r['ir'], color_mid_r['ir']) + generate_result(mid_r['ti'], color_mid_r['ti']) + generate_result(mid_r['up'], color_mid_r['up']) + "\n\t</tr>"
    bot = "\n\t<tr>\n\t\t<td>" + bot_a + "</td>" + generate_result(bot_r['gl'], color_bot_r['gl']) + generate_result(bot_r['ir'], color_bot_r['ir']) + generate_result(bot_r['ti'], color_bot_r['ti']) + generate_result(bot_r['up'], color_bot_r['up']) + "\n\t</tr>"

    if tags == "":
        return '\n<table class="table-responsive-sm">{}{}{}{}{}\n</table>'.format(question, header, top, mid, bot)
    else:
        return '\n<table class="table-responsive-sm {}">{}{}{}{}{}\n</table>'.format(tags, question, header, top, mid, bot)

def filter_list(unique_shadows):
    output = '\t\t\t\t<button class="btn filter-btn active" style="font-weight:bold;" onclick="filterByShadows(\'all\')"> Show all</button>\n\t\t\t\t<button class="btn filter-btn" style="font-weight:bold;" onclick="filterByShadows(\'None\')"> Uncategorized</button>'
    for shadow in sorted(unique_shadows):
        tag = shadow.replace(' ', '_')
        btn = '\n\t\t\t\t<button class="btn filter-btn" onclick="filterByShadows(\'{}\')"> {}</button>'.format(tag, shadow)
        output += btn
    return output

def convert_csv_to_new_tables(datafilename, outputfilename, indexfilename, royalFlag):
    if royalFlag:
        print("Generating ROYAL Tables...")
    else:
        print("Generating ORIGINAL Tables...")

    wb = load_workbook(datafilename)
    ws = wb.active

    outputfile = open(outputfilename, 'w', encoding="utf8")
    indexfile = open(indexfilename, encoding="utf8")
    
    table_text = []
    unique_shadows = []
    all_shadow_groups = []

    for r, row in enumerate(ws.values):
        shadows = row[1]
        question_text = row[2]
        first_text = row[4]
        first_res = {
            'gl': row[5],
            'ir': row[6],
            'ti': row[7],
            'up': row[8]
        }
        color_first_res = {
            'gl': False if ws.cell(row=r+1, column=6).fill.start_color.index == 'FFFFFF00' or row[5] == '-' else True,
            'ir': False if ws.cell(row=r+1, column=7).fill.start_color.index == 'FFFFFF00' or row[6] == '-' else True,
            'ti': False if ws.cell(row=r+1, column=8).fill.start_color.index == 'FFFFFF00' or row[7] == '-' else True,
            'up': False if ws.cell(row=r+1, column=9).fill.start_color.index == 'FFFFFF00' or row[8] == '-' else True
        }
        second_text = row[9]
        second_res = {
            'gl': row[10],
            'ir': row[11],
            'ti': row[12],
            'up': row[13]
        }
        color_second_res = {
            'gl': False if ws.cell(row=r+1, column=11).fill.start_color.index == 'FFFFFF00' or row[10] == '-' else True,
            'ir': False if ws.cell(row=r+1, column=12).fill.start_color.index == 'FFFFFF00' or row[11] == '-' else True,
            'ti': False if ws.cell(row=r+1, column=13).fill.start_color.index == 'FFFFFF00' or row[12] == '-' else True,
            'up': False if ws.cell(row=r+1, column=14).fill.start_color.index == 'FFFFFF00' or row[13] == '-' else True
        }
        third_text = row[14]
        third_res = {
            'gl': row[15],
            'ir': row[16],
            'ti': row[17],
            'up': row[18]
        }
        color_third_res = {
            'gl': False if ws.cell(row=r+1, column=16).fill.start_color.index == 'FFFFFF00' or row[15] == '-' else True,
            'ir': False if ws.cell(row=r+1, column=17).fill.start_color.index == 'FFFFFF00' or row[16] == '-' else True,
            'ti': False if ws.cell(row=r+1, column=18).fill.start_color.index == 'FFFFFF00' or row[17] == '-' else True,
            'up': False if ws.cell(row=r+1, column=19).fill.start_color.index == 'FFFFFF00' or row[18] == '-' else True
        }

        if shadows != None and r != 0:
            tags = shadows.split(', ')
            for t, tag in enumerate(tags):
                if tag not in unique_shadows and tag != " ":
                    unique_shadows.append(tag)
                tags[t] = tag.replace(' ', '_')
            tags = 'filterDiv ' + ' '.join(tags)
        else:
            tags = 'filterDiv None'
        
        if shadows != None and r != 0:
            new_tags = shadows.split(', ')
            for t, new_tag in enumerate(new_tags):
                new_tags[t] = new_tag.replace(' ', '_')
            all_shadow_groups.append(new_tags)

        if r != 0:
            table_text.append(create_table(question_text, tags, first_text, second_text, third_text, first_res, second_res, third_res, color_first_res, color_second_res, color_third_res))

    indexHTML = indexfile.read().split('<div id="questions">')
    top_text = indexHTML[0]
    top_text = re.sub(r'<div class="card-body" id="shadowFilterBtns">(\s*<button.*)*', '<div class="card-body" id="shadowFilterBtns">\n' + filter_list(unique_shadows), top_text)

    outputfile.write(top_text)
    outputfile.write('<div id="questions">')
    for table in table_text:
        outputfile.write(table)
    outputfile.write("\n</div>")

    indexHTML_script = indexHTML[1].split('<!-- LOCAL STORAGE OF SWITCH STATE -->')
    bot_text = indexHTML_script[1]
    outputfile.write('\n\n<!-- LOCAL STORAGE OF SWITCH STATE -->')
    outputfile.write(bot_text)

    outputfile.close()

    Output = list(merge_common(all_shadow_groups)) 

    if royalFlag:
        print("---- Finished ROYAL Tables ----")
        # pprint(Output)
    else:
        print("---- Finished ORIGINAL Tables ----")
        # pprint(Output)


royal_file = 'data/persona-5-royal-questions.xlsx'
royal_output = 'data/output-royal.html'
royal_index = 'royal.md'

original_file = 'data/persona-5-questions.xlsx'
original_output = 'data/output-original.html'
original_index = 'index.md'

convert_csv_to_new_tables(royal_file, royal_output, royal_index, True)
convert_csv_to_new_tables(original_file, original_output, original_index, False)
